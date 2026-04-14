import os
from datetime import datetime, timezone
from typing import List, Dict, Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.logger import logger

COLUMNS = [
    "post_id",
    "creator",
    "post_url",
    "caption",
    "likes",
    "comments",
    "views",
    "video_url",
    "audio_url",
    "video_duration",
    "image_url",
    "posted_at",
    "fetched_at",
]

HEADER_FILL   = PatternFill("solid", start_color="2E4057")   # dark blue
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CELL_FONT     = Font(name="Arial", size=10)
ALT_FILL      = PatternFill("solid", start_color="F0F4F8")   # light grey alternate rows
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

COL_WIDTHS = {
    "post_id":        28,
    "creator":        22,
    "post_url":       45,
    "caption":        60,
    "likes":          10,
    "comments":       10,
    "views":          10,
    "video_url":      45,
    "audio_url":      40,
    "video_duration": 16,
    "image_url":      45,
    "posted_at":      22,
    "fetched_at":     22,
}


def _get_or_create_workbook(path: str) -> openpyxl.Workbook:
    if os.path.exists(path):
        return openpyxl.load_workbook(path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)      # remove default blank sheet
    return wb


def _get_or_create_sheet(wb: openpyxl.Workbook, creator: str) -> openpyxl.worksheet.worksheet.Worksheet:
    sheet_name = creator[:31]   # Excel sheet name max 31 chars
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.font   = HEADER_FONT
        cell.fill   = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 20)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    return ws


def _existing_post_ids(ws) -> set:
    ids = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            ids.add(str(row[0]))
    return ids


def _append_row(ws, post: Dict[str, Any], row_num: int):
    values = [
        post.get("external_post_id", ""),
        post.get("creator_id", ""),
        post.get("post_url", ""),
        post.get("text", ""),
        post.get("likes", 0),
        post.get("comments", 0),
        post.get("views", 0),
        post.get("video_url", ""),
        post.get("audio_url", ""),
        post.get("video_duration", ""),
        post.get("image_url", ""),
        post.get("timestamp", ""),
        datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
    ]

    is_alt = (row_num % 2 == 0)

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = CELL_FONT
        cell.alignment = LEFT if col_idx in (4,) else CENTER   # caption left-aligned
        if is_alt:
            cell.fill = ALT_FILL

    ws.row_dimensions[row_num].height = 18


def save_posts_to_excel(posts: List[Dict[str, Any]], excel_path: str) -> Dict[str, int]:
    """
    Upserts posts into an Excel file, one sheet per creator.
    Skips posts whose post_id already exists in the sheet.
    Returns counts: {creator: new_posts_added}
    """
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    wb = _get_or_create_workbook(excel_path)

    stats: Dict[str, int] = {}

    # Group posts by creator
    by_creator: Dict[str, List[Dict]] = {}
    for post in posts:
        creator = post.get("creator_id", "unknown")
        by_creator.setdefault(creator, []).append(post)

    for creator, creator_posts in by_creator.items():
        ws = _get_or_create_sheet(wb, creator)
        existing_ids = _existing_post_ids(ws)
        next_row = ws.max_row + 1
        added = 0

        for post in creator_posts:
            post_id = str(post.get("external_post_id", ""))
            if not post_id or post_id in existing_ids:
                continue
            _append_row(ws, post, next_row)
            existing_ids.add(post_id)
            next_row += 1
            added += 1

        stats[creator] = added
        logger.info(f"Excel storage: {added} new posts added for creator '{creator}'")

    wb.save(excel_path)
    logger.info(f"Excel file saved → {excel_path}")
    return stats
